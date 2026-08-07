"""Tests notifier Telegram + export Excel (spec MISSION §8)."""
from datetime import datetime, timezone

import httpx
import pytest
import respx
from openpyxl import load_workbook

from surebet.ai.scout import Scout
from surebet.export.excel import COLUMNS, export_opportunities
from surebet.normalizer.schema import Odd
from surebet.notifier.telegram import TelegramNotifier, format_alert, should_alert


def _odd(bookmaker, selection, odds):
    return Odd(
        bookmaker=bookmaker, sport="football", competition="Amical", match_id="m1",
        home_team="Ghana", away_team="Colombie",
        start_time=datetime(2026, 7, 23, 18, 0, tzinfo=timezone.utc),
        market_type="1x2", n_outcomes=3, selection=selection, line=None, team_scope=None,
        odds=odds, url=f"https://{bookmaker}.example/bet", scraped_at=datetime.now(timezone.utc),
    )


def _three_way_opp(score=88):
    scout = Scout(bankroll=50_000.0)
    opp = scout.evaluate([
        _odd("Paryaj Lakay", "home", 3.55),
        _odd("1xBet", "draw", 3.90),
        _odd("Golcash", "away", 3.30),
    ])[0]
    opp.score_ia = score
    opp.explanation = "Fè vit!"
    return opp


class TestShouldAlert:
    def test_alerts_when_roi_and_score_high(self):
        assert should_alert(_three_way_opp(score=88)) is True

    def test_no_alert_when_score_below_threshold(self):
        assert should_alert(_three_way_opp(score=60)) is False

    def test_no_alert_when_score_missing(self):
        opp = _three_way_opp()
        opp.score_ia = None
        assert should_alert(opp) is False

    def test_no_alert_when_roi_below_2pct(self):
        scout = Scout(bankroll=50_000.0)
        # M=0.98 -> ROI ~2.04% ; forcons un ROI < 2 via cotes serrees
        opp = scout.evaluate([
            _odd("A", "home", 3.30),
            _odd("B", "draw", 3.30),
            _odd("C", "away", 3.30),
        ])[0]
        opp.score_ia = 90
        # ROI ~ 1.0% (M ~ 0.909... wait 3/3.3=0.909 -> ROI 10%) -> ensure logic uses threshold
        # Ici ROI eleve ; on teste la borne avec un cas construit:
        opp.roi_pct = 1.5
        assert should_alert(opp) is False


def test_format_alert_contains_stakes_links_and_explanation():
    opp = _three_way_opp()
    text = format_alert(opp)
    assert "SUREBET" in text
    assert "Paryaj Lakay" in text and "1xBet" in text and "Golcash" in text
    assert "https://Paryaj Lakay.example/bet" in text or "Placer le pari" in text
    assert "Profit garanti" in text
    assert "Fè vit!" in text


def test_format_alert_is_valid_html_with_special_chars():
    """Regression : market_type "shots_team" (underscore) et noms d'equipes
    cassaient le parseur Markdown herite -> on emet du HTML echappe (parse_mode
    HTML), teste envoye avec succes en live sur Telegram."""
    opp = _three_way_opp()
    opp.market_type = "shots_team"          # underscore : cassait le Markdown
    opp.match_label = "A & B <C>"           # &, <, > : doivent etre echappes
    text = format_alert(opp)
    # balises HTML valides, pas de Markdown
    assert "<b>SUREBET" in text and "</b>" in text
    assert "<a href=" in text
    # caracteres speciaux echappes
    assert "A &amp; B &lt;C&gt;" in text
    # l'underscore n'est plus un marqueur d'italique
    assert "shots_team" in text


@pytest.mark.asyncio
@respx.mock
async def test_notifier_uses_html_parse_mode():
    import json as _json

    captured = {}

    def _capture(request):
        captured.update(_json.loads(request.content))
        return httpx.Response(200, json={"ok": True})

    respx.post(url__regex=r".*/sendMessage").mock(side_effect=_capture)
    notifier = TelegramNotifier(bot_token="123:abc", chat_id="@chan")
    await notifier.send(_three_way_opp())
    assert captured["parse_mode"] == "HTML"


@pytest.mark.asyncio
async def test_notifier_not_configured_returns_false():
    notifier = TelegramNotifier(bot_token=None, chat_id=None)
    assert notifier.is_configured is False
    assert await notifier.send(_three_way_opp()) is False


@pytest.mark.asyncio
@respx.mock
async def test_notifier_sends_when_configured():
    route = respx.post(url__regex=r"https://api\.telegram\.org/bot.*/sendMessage").mock(
        return_value=httpx.Response(200, json={"ok": True})
    )
    notifier = TelegramNotifier(bot_token="123:abc", chat_id="999")
    assert await notifier.send(_three_way_opp()) is True
    assert route.called


def test_excel_export_columns_and_rows(tmp_path):
    opp = _three_way_opp()
    out = export_opportunities([opp], tmp_path / "surebets.xlsx")
    wb = load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == COLUMNS
    row2 = [c.value for c in ws[2]]
    assert row2[6] == "football"  # SPORT
    assert row2[7] == "Ghana - Colombie"  # MATCH
    assert row2[8] == "Paryaj Lakay"  # BOOKMAKER A
    assert row2[16] == "Golcash"  # BOOKMAKER C

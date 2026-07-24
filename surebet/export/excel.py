"""Export Excel reproduisant les colonnes du fichier fourni (spec MISSION §8)."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from ..normalizer.schema import Opportunity

# Colonnes exactes du fichier source (spec MISSION §8)
COLUMNS = [
    "BET DATE", "MATCH DATE", "INIT BALANCE", "FINAL BALANCE", "% SUREBET", "PROFIT",
    "SPORT", "MATCH",
    "BOOKMAKER A", "EVENT", "ODDS A", "MISE",
    "BOOKMAKER B", "EVENT", "ODDS B", "MISE",
    "BOOKMAKER C", "EVENT", "ODDS C", "MISE",
]


def _row_from_opportunity(opp: Opportunity) -> list:
    legs = opp.legs
    a = legs[0]
    b = legs[1]
    c = legs[2] if len(legs) > 2 else None
    return [
        opp.detected_at.strftime("%Y-%m-%d %H:%M"),
        opp.match_date.strftime("%Y-%m-%d %H:%M") if opp.match_date else "",
        round(opp.bankroll, 2),
        round(opp.bankroll + opp.profit, 2),
        round(opp.roi_pct, 2),
        round(opp.profit, 2),
        opp.sport,
        opp.match_label,
        a.bookmaker, a.event_label, a.odds, round(a.stake, 2),
        b.bookmaker, b.event_label, b.odds, round(b.stake, 2),
        c.bookmaker if c else "", c.event_label if c else "", c.odds if c else "", round(c.stake, 2) if c else "",
    ]


def export_opportunities(opportunities: list[Opportunity], path: str | Path) -> Path:
    """Ecrit un classeur .xlsx avec les colonnes exactes du fichier fourni."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Surebets"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for opp in opportunities:
        ws.append(_row_from_opportunity(opp))

    for column_cells in ws.columns:
        width = max((len(str(c.value)) for c in column_cells if c.value is not None), default=10)
        ws.column_dimensions[column_cells[0].column_letter].width = min(width + 2, 40)

    path = Path(path)
    wb.save(path)
    return path
